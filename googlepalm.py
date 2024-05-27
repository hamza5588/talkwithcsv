from django.shortcuts import render
import pandas as pd
import numpy as np
from llama_index.core.tools import FunctionTool
import nest_asyncio
from llama_index.llms.anthropic import Anthropic
from llama_parse import LlamaParse
from llama_index.llms.openai import OpenAI
from llama_index.core.tools import QueryEngineTool, ToolMetadata
import joblib
import io
import re
import xlrd
import xlwt
from llama_index.core.agent import FunctionCallingAgentWorker
from llama_index.core.node_parser import SentenceSplitter

from llama_index.llms.openrouter import OpenRouter
from django.http import HttpResponse
from langchain.agents.agent_types import AgentType
from langchain_experimental.agents.agent_toolkits import create_pandas_dataframe_agent

from langchain.llms.base import BaseLLM
from llama_index.llms.openrouter import OpenRouter
from langchain_openai import ChatOpenAI
import mimetypes
import pandas as pd
from langchain_openai import OpenAI
import os
import shutil


from django.conf import settings

from django.shortcuts import redirect

from questionapp import settings
nest_asyncio.apply()
import os
import csv
from django.http import JsonResponse


import openpyxl


import os
from django.http import HttpResponse, Http404

import requests

import openai

import os

from django.http import FileResponse
import pandas as pd
import os
from pandasai import PandasAI
from pandasai.llm.google_palm import GooglePalm



from django.conf import settings
import os




from llama_index.embeddings.openai import OpenAIEmbedding

from llama_index.core import Settings
# from llama_index.embeddings.huggingface import HuggingFaceEmbedding

apikey="sk-proj-vHGiY1djyz184bFNj9ZUT3BlbkFJlWslQikn2yUZq8pP4Dni"
llmaparsekey="llx-vr8V8896QhMSPiR9tsWmyQkqAzDZVoVKQ4wPP0k89wsKG0AD"
cloudapi="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"
openrouterapi = "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"
googleapikey="AIzaSyAAcMvFrSK26pDT84-P0nYVGGS4_4bJjHU"

# class OpenRouterAI(BaseLLM):
#     def __init__(self, api_key: str, **kwargs):
#         super().__init__(**kwargs)
#         self.api_key = api_key
#         self.client = OpenRouterClient(api_key)

#     def _call(self, prompt: str, **kwargs) -> str:
#         response = self.client.generate_text(prompt)
#         return response

#     @property
#     def _identifying_params(self) -> Dict[str, Any]:
#         return {"api_key": self.api_key}



def readfile(a):
    file_extension = a.split('.')[-1]  # Get the file extension
    
    if file_extension == 'csv':
        return pd.read_csv(a)
    elif file_extension == 'xlsx':
        return pd.read_excel(a)
    else:
        raise ValueError("Unsupported file format. Only CSV and Excel files are supported.")
    



uploaded_data=None

def upload_file(request):
    global uploaded_data
    global file_extension
    global file_name
    if request.method == 'POST':
        uploaded_file = request.FILES['file']
        file_name, file_extension = os.path.splitext(uploaded_file.name)

        if file_extension.lower() == '.csv':
            # Handle CSV file
            data = pd.read_csv(uploaded_file, encoding='utf-8')
            uploaded_data=data
        elif file_extension.lower() in ['.xls', '.xlsx']:
            # Handle Excel file
            data = pd.read_excel(uploaded_file)
            print(data)

            uploaded_data=data

            # Create a directory to store the converted CSV file if it doesn't exist
            csv_dir = os.path.join(settings.MEDIA_ROOT, 'uploaded_files')
            if not os.path.exists(csv_dir):
                os.makedirs(csv_dir)

           

            # Save the data as a CSV file in the directory
            csv_file_path = os.path.join(csv_dir, 'data.csv')
            data.to_csv(csv_file_path, index=False)

            # Read the saved CSV file
            data = pd.read_csv(csv_file_path, encoding="utf-8")
            print(data)
        else:
            # Unsupported file type
            return HttpResponse("Unsupported file type. Please upload a CSV or Excel file.")

        # Create a directory to store the uploaded file if it doesn't exist
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploaded_files')
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        # Move the uploaded file to the directory
        file_path = os.path.join(upload_dir, uploaded_file.name)
        with open(file_path, 'wb') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
      
        return redirect("resultfile")
    else:
        return render(request,"chat.html")
    
def downloadfile(request):
    # Get the file name and extension from the request
    global file_extension
    global file_name_with_extension
    file_name_with_extension = file_name + file_extension

    # Construct the file path using the MEDIA_ROOT and the file name with extension
    file_path = os.path.join(settings.MEDIA_ROOT, 'download', file_name_with_extension)

    # Check if the file exists before proceeding
    if os.path.exists(file_path):
        # Return the file as a response
        if file_extension in ['.xls', '.xlsx']:
            response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name_with_extension, content_type='application/vnd.ms-excel')
        else:
            response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name_with_extension)
        return response
    else:
        # If the file doesn't exist, return an appropriate error message
        return HttpResponse('File not found.', status=404)




def check_response(response):
    # Define a list of common Pandas function names
    pandas_functions = ['loc', 'iloc', 'apply', 'groupby', 'merge', 'sort_values']
    
    # Check for Pandas function names
    for func in pandas_functions:
        if f'df.{func}' in response.lower():
            return False
    
    # Check for column access syntax
    if re.search(r"df\['[\w\s]+'\]", response.lower()):
        return False
    
    # Check for data types
    if any(dtype in response.lower() for dtype in ['int64', 'float64', 'object']):
        return False
    
    # Check for code formatting (indentation)
    if any(line.startswith(' ') for line in response.split('\n')):
        return False
    
    # If none of the patterns are found, assume it's a human-friendly response
    return True







def result(request):
    if request.method == 'POST':
        global uploaded_data
        global file_name
        file = request.POST.get("names")

        #googlepalm
    
      
        df = uploaded_data
        llm = GooglePalm(api_key=googleapikey)
        pandas_ai = PandasAI(llm)
        pandas_ai.run(df, prompt=file)
        print(df)
        
        download_dir = os.path.join(settings.MEDIA_ROOT, 'download')
        os.makedirs(download_dir, exist_ok=True)

        # Construct the file path
        ufile=file_name+file_extension
        filepath = os.path.join(download_dir, f'{ufile}')

        if file_extension.lower() == '.csv':
            df.to_csv(filepath, index=False)
        elif file_extension.lower() == '.xlsx':
            df.to_excel(filepath, index=False)

        else:
            
             xlsx_file = os.path.splitext(filepath)[0] + '.xlsx'
             df.to_excel(xlsx_file, index=False)
             workbook = openpyxl.load_workbook(xlsx_file)
             xls_file = os.path.splitext(filepath)[0] + '.xls'
             writer = xlwt.Workbook()
             for sheet_name in workbook.sheetnames:
                 worksheet = workbook[sheet_name]
                 writer_sheet = writer.add_sheet(sheet_name)
                 for row in worksheet.iter_rows():
                     for cell in row:
                         writer_sheet.write(cell.row - 1, cell.column - 1, cell.value)
             writer.save(xls_file)


        

        
        ufile=file_name+file_extension

        filepath = os.path.join(settings.MEDIA_ROOT, 'download', f'{ufile}')
        print(filepath)

        if file_extension.lower() == '.csv':
            with open(filepath, 'r',encoding='utf-8', newline='') as file:
                csv_reader = csv.reader(file)
                csv_data = list(csv_reader)
        

        elif file_extension.lower() in ['.xls', '.xlsx']:
            xls_data = pd.read_excel(filepath)
            csv_data = xls_data.values.tolist()
            csv_data.insert(0, list(xls_data.columns)) 

        # # Read the CSV file and store its content in a list of lists
        # with open(filepath, 'r',encoding='utf-8', newline='') as file:
        #     csv_reader = csv.reader(file)
        #     csv_data = list(csv_reader)


     

        filepath=r"C:\Users\PMLS\Desktop\project\questionapp\file.csv"

        b="Succussfull"


        
     
        
        return render(request, 'result.html',locals())
    else:
        return render(request, "result.html")
